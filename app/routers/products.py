from decimal import Decimal
from io import BytesIO
from datetime import datetime, date

from fastapi import Query
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.deps import require_shop_owner
from app.core.discounts import calculate_discount
from app.db.session import get_db
from app.models.product import Product
from app.models.shop import Shop
from app.models.user import User
from app.schemas.product import ProductCreate, ProductRead, ProductUpdate
from app.schemas.product_public import ProductRead as ProductPublicRead


router = APIRouter(prefix="/products", tags=["products"])
shop_products_router = APIRouter(tags=["products"])

def _get_owner_shop(db: Session, owner_id: int) -> Shop:
    shop = db.scalar(select(Shop).where(Shop.owner_id == owner_id))
    if not shop:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Shop not found. Register your shop first.",
        )
    return shop


@router.get("", response_model=list[ProductRead])
def list_products(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_shop_owner),
    limit: int = Query(10, le = 100),
    offset: int = Query(0),
) -> list[Product]:
    shop = _get_owner_shop(db, current_user.id)
    stmt = select(Product).where(Product.shop_id == shop.id).order_by(Product.id.desc()).offset(offset).limit(limit)
    return list(db.scalars(stmt).all())


@router.post("", response_model=ProductRead, status_code=status.HTTP_201_CREATED)
def create_product(
    payload: ProductCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_shop_owner),
) -> Product:
    shop = _get_owner_shop(db, current_user.id)
    if payload.shop_id is not None and payload.shop_id != shop.id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You can only add products to your own shop",
        )

    discount_percent, final_price = calculate_discount(payload.expiry_date, payload.price)

    product = Product(
        shop_id=shop.id,
        name=payload.name,
        description=payload.description,
        image_url=payload.image_url,
        price=payload.price,
        stock=payload.stock,
        expiry_date=payload.expiry_date,
        discount_percent=discount_percent,
        final_price=final_price,
    )
    db.add(product)
    db.commit()
    db.refresh(product)
    return product


@router.put("/{product_id}", response_model=ProductRead)
def update_product(
    product_id: int,
    payload: ProductUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_shop_owner),
) -> Product:
    shop = _get_owner_shop(db, current_user.id)
    product = db.get(Product, product_id)
    if not product or product.shop_id != shop.id:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Product not found"
        )

    if payload.name is not None:
        product.name = payload.name
    if payload.description is not None:
        product.description = payload.description
    if payload.image_url is not None:
        product.image_url = payload.image_url
    if payload.price is not None:
        product.price = payload.price
    if payload.stock is not None:
        product.stock = payload.stock
    if payload.expiry_date is not None:
        product.expiry_date = payload.expiry_date
    price = Decimal(str(product.price))
    discount_percent, final_price = calculate_discount(product.expiry_date, price)
    product.discount_percent = discount_percent
    product.final_price = final_price

    db.add(product)
    db.commit()
    db.refresh(product)
    return product


@router.post("/bulk-upload", status_code=status.HTTP_201_CREATED)
async def bulk_upload_products(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_shop_owner),
) -> dict:
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only Excel files are supported",
        )

    shop = _get_owner_shop(db, current_user.id)
    content = await file.read()

    try:
        workbook = load_workbook(filename=BytesIO(content), data_only=True)
        sheet = workbook.active
        if sheet is None or not isinstance(sheet, Worksheet):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Excel file has no valid active worksheet",
            )
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid Excel file: {exc}",
        )

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail="Excel file is empty"
        )

    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    index = {h: i for i, h in enumerate(headers)}
    required = {"name", "price", "stock"}
    missing = sorted(required - set(index.keys()))
    if missing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Missing required columns: {missing}",
        )

    created = 0
    errors: list[dict] = []
    for row_num, row in enumerate(rows[1:], start=2):
        try:
            name = str(row[index["name"]]).strip() if row[index["name"]] is not None else ""
            if not name:
                raise ValueError("name is required")

            def safe_int(value) -> int:
                if value is None:
                    raise ValueError("Missing integer value")
                return int(float(value))
            
            def safe_decimal(value) -> Decimal:
                if value is None:
                    raise ValueError("Missing decimal value..")
                return Decimal(str(value))
        
            price = safe_decimal(row[index["price"]])
            stock = safe_int(row[index["stock"]])

            description = (
                str(row[index["description"]]).strip()
                if "description" in index and row[index["description"]] is not None
                else None
            )
            image_url = (
                str(row[index["image_url"]]).strip()
                if "image_url" in index and row[index["image_url"]] is not None
                else None
            )
            expiry_date = (
                row[index["expiry_date"]]
                if "expiry_date" in index and row[index["expiry_date"]] is not None
                else None
            )
            if isinstance(expiry_date, datetime):
                expiry_date = expiry_date.date()
            elif not isinstance(expiry_date, date):
                expiry_date = None

            discount_percent, final_price = calculate_discount(expiry_date, price)
            db.add(
                Product(
                    shop_id=shop.id,
                    name=name,
                    description=description,
                    image_url=image_url,
                    price=price,
                    stock=stock,
                    expiry_date=expiry_date,
                    discount_percent=discount_percent,
                    final_price=final_price,
                )
            )
            created += 1
        except Exception as exc:
            errors.append({"row": row_num, "error": str(exc)})

    db.commit()
    return {"created_count": created, "error_count": len(errors), "errors": errors}


@router.delete("/{product_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_product(
    product_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_shop_owner),
) -> None:
    shop = _get_owner_shop(db, current_user.id)
    product = db.get(Product, product_id)
    if not product or product.shop_id != shop.id:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Product not found"
        )
    db.delete(product)
    db.commit()
    return None


@router.get("/{product_id}", response_model=ProductRead)
def get_product(
    product_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_shop_owner),
) -> Product:
    shop = _get_owner_shop(db, current_user.id)
    product = db.get(Product, product_id)
    if not product or product.shop_id != shop.id:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Product not found"
        )
    return product


@shop_products_router.get("/shops/{shop_id}/products", response_model=list[ProductPublicRead])
def list_shop_products(shop_id: int, db: Session = Depends(get_db)) -> list[ProductPublicRead]:
    from datetime import date

    today = date.today()

    stmt = (
        select(Product)
        .where(
            Product.shop_id == shop_id,
            Product.stock > 0,
            (Product.expiry_date.is_(None) | (Product.expiry_date >= today)),
        )
        .order_by(Product.id.desc())
    )
    products = list(db.scalars(stmt).all())

    result: list[ProductPublicRead] = []
    for product in products:
        price = Decimal(str(product.price))
        discount_percent, final_price = calculate_discount(product.expiry_date, price)
        result.append(
            ProductPublicRead(
                id=product.id,
                name=product.name,
                description=product.description,
                price=price,
                discount_percent=discount_percent,
                final_price=final_price,
                expiry_date=product.expiry_date,
                stock=product.stock,
                image_url=product.image_url,
            )
        )

    return result

